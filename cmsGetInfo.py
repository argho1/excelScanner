import os
import openpyxl
import pandas as pd

# Function to normalize a list of values (lowercase and remove dots, spaces, and non-breaking spaces)
def normalize_values(values):
    normalized_values = [str(value).lower().replace(' ', '').replace('.', '').replace('\xa0', '') for value in values]
    return normalized_values

# Function to extract data below the search values in the same column
def extract_data_below_values(sheet, search_values):
    # Iterate through all cells and search for the specified values
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            # Normalize the cell value for comparison
            normalized_cell_value = str(cell.value).lower().replace(' ', '').replace('.', '').replace('\xa0', '')

            # Search for the normalized values
            if any(search_value in normalized_cell_value for search_value in search_values):
                # Get the row and column indices
                row_index = cell.row
                col_index = cell.column

                # Extract the data below the values in the same column
                data = [sheet.cell(i, col_index).value for i in range(row_index + 1, sheet.max_row + 1)]

                # Return a tuple with both data and column index
                return pd.DataFrame({f"{sheet.title}_{col_index}": data}), col_index

    # If none of the search values are found, extract all non-empty values from the first column
    first_column_data = [sheet.cell(i, 1).value for i in range(1, sheet.max_row + 1) if sheet.cell(i, 1).value is not None]

    # Return a DataFrame with the first column data and None for the column index
    return pd.DataFrame({f"{sheet.title}_1": first_column_data}), None


def add_numerical_values(strings):
    total = 0
    for date_string in strings:
        # Split the string at the underscore
        parts = date_string.split('_')
        if len(parts) == 2 and parts[0].isdigit():
            # Filter numerical values before the underscore
            numerical_value = int(parts[0])
            total += numerical_value
    return total

# Function to find values in an Excel file
def find_SN_in_excel(excel_file_path, search_values):
    # Load the workbook
    workbook = openpyxl.load_workbook(excel_file_path)

    # List to store DataFrames with the requested data
    data_frames_with_values = []

    all_values = []
    all_sheet_names = []

    # Iterate through each sheet
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        all_sheet_names.append(sheet_name)

        # Extract data below the search values in the same column
        extracted_data, col_index = extract_data_below_values(sheet, search_values)

        # Print the result
        if not extracted_data.empty:
            # Drop rows with NaN values
            if col_index is not None:
                extracted_data = extracted_data.dropna(subset=[f"{sheet.title}_{col_index}"])

            # Reset index to remove the default integer index
            extracted_data = extracted_data.reset_index(drop=True)

            # Append all non-null values to the list
            all_values.extend(extracted_data.values.flatten().tolist())

    # Remove any remaining non-breaking space characters
    #all_values = [value.replace('\xa0', '') for value in all_values]

    # Get no. from sheet name and add
    total_sn_count = add_numerical_values(all_sheet_names)

    print("Total SN according to sheet name : ",total_sn_count)

    # Return the list of all values found
    return all_values


def traverse_folder(folder_path):
    excel_files = []
    other_files = []

    for root, dirs, files in os.walk(folder_path):

        for file_name in files:
            file_path = os.path.join(root, file_name)

            # Check if the file has a ".xlsx" extension
            if file_name.lower().endswith(".xlsx"):
                excel_files.append(file_path)
            else:
                other_files.append(file_path)

    return excel_files, other_files






# Specify the search values (normalize them)
search_values = normalize_values(['router sn', 'serial no', 'router sr no', 'rputer sn'])
'''
# Specify the path to the Excel file
excel_file_path = '../customerSNexcel/Securesource.xlsx'
# Call the function to find values in the Excel file
customerSN_list = find_SN_in_excel(excel_file_path, search_values)
'''




cmsSN_list = find_SN_in_excel('../cms_router_list.xlsx', search_values)

# Folder to traverse
excel_files_path_list, other_files_path_list = traverse_folder('../customerSNexcel/dd')


for excel_files_path in excel_files_path_list:
    
    
    print()
    excel_file_name = os.path.basename(excel_files_path)

    if excel_file_name.startswith("~$"):
        print(f"\nSlipping file {excel_file_name}\n")
    
    else:
        print(f'Searching file {excel_file_name}')
        customerSN_list = find_SN_in_excel(excel_files_path, search_values)
        print("Customer SN Found : ", len(customerSN_list))

        # Print the list of all values found
        
        print()
        print("Total SN in CMS : ", len(cmsSN_list))
        


        # Find the common elements (duplicates) between the two lists
        common_elements = list(set(customerSN_list) & set(cmsSN_list))

        # Print the list of duplicate values
        print("Customer SN in CMS : ", len(common_elements))
        print()
        print()






'''
print(excel_files_list)
print()
print(other_files_list)
'''